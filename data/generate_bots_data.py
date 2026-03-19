"""
generate_bots_data.py
---------------------
Generates data/dreamco_bots_data.xlsx (interactive Excel workbook with
AutoFilter, frozen header row, and colour-coded ROI priority) and a plain
data/dreamco_bots_data.csv companion file from the master bot catalog.

Usage
-----
    python data/generate_bots_data.py

Output
------
    data/dreamco_bots_data.xlsx
    data/dreamco_bots_data.csv
"""

import csv
import os
from pathlib import Path

# ---------------------------------------------------------------------------
# Master bot catalog
# Each entry: (Bot Name, Slug, Price USD/mo, Industry, Function, ROI Priority,
#              Description)
# ROI Priority: "High" / "Medium" / "Low"
# ---------------------------------------------------------------------------
BOTS = [
    # ── Finance & Insurance ────────────────────────────────────────────────
    (
        "AI Credit Risk Underwriter",
        "credit-underwriter",
        999,
        "Finance",
        "Analytics",
        "High",
        "Automated credit risk assessment and underwriting with explainable AI decisions.",
    ),
    (
        "Capital Allocation Optimizer",
        "capital-alloc-optimizer",
        799,
        "Finance",
        "Forecasting",
        "High",
        "Optimizes capital allocation with risk-adjusted returns across portfolio investments.",
    ),
    (
        "Cash Flow Stress-Test AI",
        "cashflow-stress-test",
        599,
        "Finance",
        "Analytics",
        "High",
        "Stress-tests cash flow under various adverse scenarios to assess financial resilience.",
    ),
    (
        "Currency Hedging Bot",
        "currency-hedging",
        499,
        "Finance",
        "Automation",
        "Medium",
        "Automated FX hedging strategies to protect international revenue streams.",
    ),
    (
        "Debt Payoff Strategist",
        "debt-payoff",
        499,
        "Finance",
        "Analytics",
        "Medium",
        "Creates debt payoff strategies using avalanche and snowball methods.",
    ),
    (
        "Insurance Claim Fraud Detector",
        "insurance-fraud",
        999,
        "Finance",
        "Analytics",
        "High",
        "AI-powered fraud detection and claims triage for insurance carriers.",
    ),
    (
        "Insurance Rate Shopper",
        "insurance-shopper",
        599,
        "Finance",
        "Operations",
        "Medium",
        "Shops insurance rates across carriers with coverage comparison and switching assistance.",
    ),
    (
        "Robo-Advisory Platform Bot",
        "robo-advisor",
        499,
        "Finance",
        "Automation",
        "High",
        "White-label robo-advisory engine for automated portfolio management.",
    ),
    (
        "Multi-Currency Revenue Optimizer",
        "multi-currency-optimizer",
        799,
        "Finance",
        "Optimization",
        "High",
        "Optimizes revenue across multiple currencies with hedging and conversion strategies.",
    ),
    (
        "Revenue Optimization Bot",
        "revenue-optimizer",
        999,
        "Finance",
        "Optimization",
        "High",
        "Optimizes revenue across all divisions with pricing and upsell strategies.",
    ),
    (
        "Revenue Tracking Engine",
        "revenue-engine",
        499,
        "Finance",
        "Analytics",
        "High",
        "Tracks revenue per bot and campaign, manages commissions and affiliate payouts.",
    ),
    (
        "Risk-Adjusted Scenario Modeler",
        "risk-scenario-modeler",
        999,
        "Finance",
        "Forecasting",
        "High",
        "Models scenarios with built-in risk adjustment for accurate decision-making.",
    ),
    (
        "Profit Per Feature Scorer",
        "profit-per-feature",
        799,
        "Finance",
        "Analytics",
        "High",
        "Scores profit contribution per feature to guide product investment decisions.",
    ),
    (
        "Enterprise Cost Consolidation Engine",
        "cost-consolidation",
        699,
        "Finance",
        "Operations",
        "High",
        "Consolidates costs across enterprise entities for unified financial visibility.",
    ),
    (
        "ESG Investment Optimizer",
        "esg-optimizer",
        499,
        "Finance",
        "Optimization",
        "Medium",
        "Optimizes portfolios for ESG compliance while maximizing risk-adjusted returns.",
    ),
    (
        "Cross-Border Payment Engine",
        "cross-border-payment",
        899,
        "Finance",
        "Automation",
        "High",
        "Automates multi-currency cross-border payments with compliance checks.",
    ),
    (
        "Algorithmic Trading Bot",
        "algo-trading",
        799,
        "Finance",
        "Automation",
        "High",
        "SMA crossover, RSI signals, and backtesting engine for algorithmic trading.",
    ),
    (
        "Loyalty Program Impact Simulator",
        "loyalty-simulator",
        499,
        "Finance",
        "Analytics",
        "Medium",
        "Models ROI, CLV uplift, and churn reduction for loyalty programs.",
    ),
    # ── Sales & Marketing ─────────────────────────────────────────────────
    (
        "Revenue-Maximizing Journey Optimizer",
        "revenue-journey-optimizer",
        699,
        "Sales & Marketing",
        "Optimization",
        "High",
        "Optimizes customer journeys to maximize revenue at every touchpoint.",
    ),
    (
        "Revenue-Impact Auto-Prioritizer",
        "revenue-impact-prioritizer",
        699,
        "Sales & Marketing",
        "Automation",
        "High",
        "Auto-prioritizes tasks and processes based on estimated revenue impact.",
    ),
    (
        "Brand Perception Steering AI",
        "brand-perception",
        699,
        "Sales & Marketing",
        "Analytics",
        "Medium",
        "Monitors and shapes brand perception across social and digital channels.",
    ),
    (
        "Customer Acquisition Cost Optimizer",
        "cac-optimizer",
        599,
        "Sales & Marketing",
        "Optimization",
        "High",
        "Reduces customer acquisition costs by optimizing ad spend and channels.",
    ),
    (
        "Blog & Article Writer",
        "blog-writer",
        299,
        "Sales & Marketing",
        "Content Creation",
        "Medium",
        "AI-powered blog and article generation with SEO optimization.",
    ),
    (
        "Brand Voice AI",
        "brand-voice",
        399,
        "Sales & Marketing",
        "Content Creation",
        "Medium",
        "Maintains consistent brand voice across all content channels.",
    ),
    (
        "DreamProduction Agency AI",
        "dream-production",
        799,
        "Sales & Marketing",
        "Content Creation",
        "High",
        "Full-service AI content production agency for video, copy, and design.",
    ),
    (
        "Multi-Source Lead Scraper",
        "multi-source-lead-scraper",
        199,
        "Sales & Marketing",
        "Automation",
        "High",
        "Scrapes leads from Google, LinkedIn, Twitter, Reddit, and Yelp.",
    ),
    # ── Operations & Automation ───────────────────────────────────────────
    (
        "Automated Task Delegation AI",
        "task-delegation",
        499,
        "Operations",
        "Automation",
        "High",
        "Intelligently delegates tasks to the right team members or bots.",
    ),
    (
        "Process Redundancy Eliminator",
        "process-redundancy",
        599,
        "Operations",
        "Optimization",
        "High",
        "Identifies and eliminates redundant business processes to reduce costs.",
    ),
    (
        "Multi-Process Coordination AI",
        "multi-process-coord",
        699,
        "Operations",
        "Automation",
        "High",
        "Coordinates multiple simultaneous business processes with conflict resolution.",
    ),
    (
        "AI Dispatch Center",
        "ai-dispatch",
        599,
        "Operations",
        "Automation",
        "High",
        "Central dispatch hub routing tasks and messages to the right bots and agents.",
    ),
    (
        "Beam Self-Learning Agent",
        "beam-self-learning",
        499,
        "Operations",
        "Automation",
        "High",
        "Automates 40+ hours/week of repetitive work through continuous self-learning.",
    ),
    (
        "Deployment Pipeline Manager",
        "deploy-pipeline",
        399,
        "Operations",
        "Automation",
        "Medium",
        "Manages CI/CD pipelines and deployment workflows for software delivery.",
    ),
    (
        "Smart Meeting Scheduler",
        "smart-meeting",
        299,
        "Operations",
        "Automation",
        "Medium",
        "Conflict detection and intelligent meeting scheduling across time zones.",
    ),
    (
        "Workplace Audit Tool",
        "workplace-audit",
        299,
        "Operations",
        "Analytics",
        "Medium",
        "5S methodology audit with automated scoring and improvement recommendations.",
    ),
    (
        "DreamCo Empire OS",
        "dreamco-empire-os",
        199,
        "Operations",
        "Automation",
        "High",
        "Command-and-control hub for the entire AI bot empire with 23 integrated modules.",
    ),
    # ── Analytics & Forecasting ───────────────────────────────────────────
    (
        "Predictive Analytics Engine",
        "predictive-analytics",
        699,
        "Analytics",
        "Forecasting",
        "High",
        "Enterprise-grade predictive analytics across sales, ops, and finance.",
    ),
    (
        "Multi-Scenario Forecasting Engine",
        "multi-scenario-forecast",
        799,
        "Analytics",
        "Forecasting",
        "High",
        "Builds and compares multiple business scenarios for strategic planning.",
    ),
    (
        "Demand Forecasting Bot",
        "demand-forecast",
        599,
        "Analytics",
        "Forecasting",
        "High",
        "AI-driven demand forecasting for inventory and supply chain optimization.",
    ),
    (
        "Data Science Agent",
        "data-science-agent",
        499,
        "Analytics",
        "Analytics",
        "High",
        "Automated data science workflows: EDA, modelling, and reporting.",
    ),
    (
        "Predictive Engagement Tool",
        "predictive-engagement",
        299,
        "Analytics",
        "Analytics",
        "Medium",
        "Scores customer engagement levels and predicts churn risk.",
    ),
    # ── Cybersecurity & Risk ──────────────────────────────────────────────
    (
        "Intrusion Anomaly Detector",
        "intrusion-detector",
        999,
        "Cybersecurity",
        "Security",
        "High",
        "Real-time network anomaly detection using ML-based intrusion analysis.",
    ),
    (
        "Rogue Bot Isolator",
        "rogue-bot-isolator",
        799,
        "Cybersecurity",
        "Security",
        "High",
        "Detects and quarantines rogue or compromised bots in the fleet.",
    ),
    (
        "Multi-Layer Defense Optimizer",
        "defense-optimizer",
        899,
        "Cybersecurity",
        "Security",
        "High",
        "Coordinates multi-layer cybersecurity defences for enterprise networks.",
    ),
    (
        "Disaster Recovery Simulator",
        "disaster-recovery",
        699,
        "Cybersecurity",
        "Simulation",
        "High",
        "Simulates disaster recovery scenarios to validate business continuity plans.",
    ),
    (
        "Privacy Compliance Bot",
        "privacy-compliance",
        599,
        "Cybersecurity",
        "Compliance",
        "High",
        "Automates GDPR, CCPA, and HIPAA privacy compliance monitoring.",
    ),
    # ── Compliance & Legal ────────────────────────────────────────────────
    (
        "Compliance & Audit Bot",
        "compliance-audit",
        799,
        "Compliance & Legal",
        "Compliance",
        "High",
        "Automated compliance auditing across regulatory frameworks.",
    ),
    (
        "Legal Risk Scorer",
        "legal-risk-scorer",
        699,
        "Compliance & Legal",
        "Analytics",
        "High",
        "Scores legal risk exposure from contracts, policies, and operations.",
    ),
    (
        "Government Contract Automation Bot",
        "gov-contract-bot",
        499,
        "Compliance & Legal",
        "Automation",
        "High",
        "Automates SAM.gov contract searches and proposal generation.",
    ),
    # ── Health & Life Sciences ─────────────────────────────────────────────
    (
        "Clinical Decision Support",
        "clinical-decision",
        999,
        "Health",
        "Analytics",
        "High",
        "AI-powered clinical decision support for evidence-based treatment recommendations.",
    ),
    (
        "Radiology Report Assistant",
        "radiology-assistant",
        899,
        "Health",
        "Analytics",
        "High",
        "Automates radiology report drafting with anomaly flagging.",
    ),
    (
        "Medical Billing Optimizer",
        "medical-billing",
        599,
        "Health",
        "Optimization",
        "High",
        "Reduces claim denials and optimizes medical billing workflows.",
    ),
    (
        "Clinical Trial Manager",
        "clinical-trial-mgr",
        999,
        "Health",
        "Operations",
        "High",
        "End-to-end clinical trial management from protocol design to reporting.",
    ),
    (
        "Mental Health Screening Bot",
        "mental-health-screen",
        299,
        "Health",
        "Analytics",
        "High",
        "PHQ-2, PHQ-9, and GAD-7 evidence-based mental health screening.",
    ),
    (
        "Drug Discovery Pipeline AI",
        "drug-discovery",
        999,
        "Health",
        "Analytics",
        "High",
        "Lipinski Rule-of-Five, ADMET prediction, and molecular docking score analysis.",
    ),
    # ── Real Estate ───────────────────────────────────────────────────────
    (
        "Blueprint Analysis AI",
        "blueprint-analysis",
        699,
        "Real Estate",
        "Analytics",
        "High",
        "Analyses architectural blueprints for cost estimation and compliance checks.",
    ),
    (
        "BIM Coordination Platform",
        "bim-coordination",
        799,
        "Real Estate",
        "Operations",
        "High",
        "Building Information Modelling coordination for multi-discipline projects.",
    ),
    (
        "Property Valuation AI",
        "property-valuation",
        599,
        "Real Estate",
        "Analytics",
        "High",
        "Automated property valuation using comparable sales and market trends.",
    ),
    (
        "Multi-Family Acquisition Bot",
        "multi-family-acq",
        499,
        "Real Estate",
        "Analytics",
        "High",
        "Analyses multi-family real estate acquisitions for ROI and cash-on-cash return.",
    ),
    (
        "Real Estate Cashflow Simulator",
        "re-cashflow-sim",
        299,
        "Real Estate",
        "Simulation",
        "High",
        "Cashflow, cap rate, CoC return, and portfolio analysis for rental properties.",
    ),
    # ── Agriculture ───────────────────────────────────────────────────────
    (
        "Agricultural Drone Scout",
        "agri-drone-scout",
        599,
        "Agriculture",
        "Operations",
        "Medium",
        "Autonomous drone scouting for crop health monitoring and pest detection.",
    ),
    (
        "Precision Agriculture Platform",
        "precision-agri",
        799,
        "Agriculture",
        "Analytics",
        "High",
        "Data-driven precision agriculture for yield optimisation and input reduction.",
    ),
    (
        "Farm Accounting Manager",
        "farm-accounting",
        299,
        "Agriculture",
        "Operations",
        "Medium",
        "Automated farm accounting, expense tracking, and subsidy management.",
    ),
    # ── Technology & Development ──────────────────────────────────────────
    (
        "API Monetization Engine",
        "api-monetization",
        399,
        "Technology",
        "Automation",
        "High",
        "Converts internal APIs into monetized products with billing and rate-limiting.",
    ),
    (
        "App Builder Bot",
        "app-builder",
        499,
        "Technology",
        "Automation",
        "High",
        "No-code app generation from natural-language specs with deployment support.",
    ),
    (
        "Software Development Bot",
        "software-dev-bot",
        699,
        "Technology",
        "Automation",
        "High",
        "Generates, reviews, and debugs code across multiple programming languages.",
    ),
    (
        "AI Level-Up Bot",
        "ai-level-up",
        399,
        "Technology",
        "Education",
        "High",
        "Gamified AI skills platform with 10 levels, token marketplace, and 101 AI companies.",
    ),
    # ── E-Commerce & Retail ───────────────────────────────────────────────
    (
        "E-Commerce Journey Optimizer",
        "ecom-journey-optimizer",
        699,
        "E-Commerce",
        "Optimization",
        "High",
        "Personalises e-commerce journeys to maximise average order value and conversion.",
    ),
    (
        "Deal Finder Bot",
        "deal-finder",
        299,
        "E-Commerce",
        "Automation",
        "Medium",
        "Finds the best deals and price drops across e-commerce platforms.",
    ),
    (
        "Affiliate Bot",
        "affiliate-bot",
        299,
        "E-Commerce",
        "Automation",
        "High",
        "Manages affiliate programmes, tracks commissions, and optimises partner networks.",
    ),
    # ── Crypto & DeFi ────────────────────────────────────────────────────
    (
        "Crypto Portfolio Bot",
        "crypto-portfolio",
        299,
        "Crypto & DeFi",
        "Analytics",
        "High",
        "Portfolio tracking, DCA strategies, and mining optimisation for 66+ coins.",
    ),
    (
        "Trade Master AI",
        "trade-master",
        699,
        "Crypto & DeFi",
        "Automation",
        "High",
        "Optimises international trade and commodity operations with AI-driven signals.",
    ),
    (
        "Open Claw Bot",
        "open-claw",
        199,
        "Crypto & DeFi",
        "Analytics",
        "Medium",
        "8 strategy types with AI ranking, simulation, NLP advisor, and client ROI tracking.",
    ),
    # ── Education ─────────────────────────────────────────────────────────
    (
        "Buddy Teach Bot",
        "buddy-teach",
        199,
        "Education",
        "Education",
        "High",
        "Multi-device broadcast engine with skill training, item detection, and personality engine.",
    ),
    (
        "AI Learning System",
        "ai-learning-system",
        299,
        "Education",
        "Education",
        "High",
        "Adaptive AI learning system that personalises curriculum to each learner.",
    ),
    (
        "Recipe Scaling Tool",
        "recipe-scaling",
        99,
        "Education",
        "Automation",
        "Low",
        "Scales recipes for any serving size with unit conversion and nutritional info.",
    ),
    (
        "Job Titles Bot",
        "job-titles-bot",
        299,
        "Education",
        "Education",
        "High",
        "Generates job descriptions, role titles, and career pathway maps for 25+ industries.",
    ),
    # ── Human Resources ───────────────────────────────────────────────────
    (
        "Selenium Job Application Bot",
        "job-application-bot",
        149,
        "Human Resources",
        "Automation",
        "High",
        "Automates job searching and applications across Indeed, LinkedIn, and Glassdoor.",
    ),
    (
        "211 Resource Eligibility Bot",
        "211-resource-bot",
        99,
        "Human Resources",
        "Operations",
        "Medium",
        "Helps users find local social services and check eligibility for assistance programmes.",
    ),
    # ── Consumer & Lifestyle ──────────────────────────────────────────────
    (
        "Money Finder Bot",
        "money-finder",
        199,
        "Consumer",
        "Analytics",
        "High",
        "Scans spending patterns to find hidden savings and unclaimed money opportunities.",
    ),
    (
        "Car Flipping Bot",
        "car-flipping",
        299,
        "Consumer",
        "Analytics",
        "High",
        "Identifies undervalued vehicles and estimates flip profit after reconditioning costs.",
    ),
    (
        "AI Side Hustle Bots",
        "ai-side-hustle",
        149,
        "Consumer",
        "Automation",
        "High",
        "AI-powered tools to identify, launch, and monetise income-generating side hustles.",
    ),
    (
        "Buddy Omniscient Bot",
        "buddy-omniscient",
        199,
        "Consumer",
        "Education",
        "High",
        "AR/VR overlays, viral challenges, omniscient knowledge engine surpassing major LLMs.",
    ),
    (
        "Buddy OS",
        "buddy-os",
        199,
        "Consumer",
        "Automation",
        "High",
        "Universal device compatibility with Bluetooth, Google Cast, AirPlay, and app framework.",
    ),
]

# ---------------------------------------------------------------------------
# Column headers
# ---------------------------------------------------------------------------
HEADERS = [
    "Bot Name",
    "Slug",
    "Price (USD/mo)",
    "Industry",
    "Function",
    "ROI Priority",
    "Description",
]

# ---------------------------------------------------------------------------
# ROI Priority colour map (openpyxl PatternFill)
# High = green, Medium = yellow, Low = red
# ---------------------------------------------------------------------------
ROI_COLORS = {
    "High": "C6EFCE",    # light green
    "Medium": "FFEB9C",  # light yellow
    "Low": "FFC7CE",     # light red/pink
}


def _write_csv(output_path: Path) -> None:
    """Write bot data to a CSV file."""
    with output_path.open("w", newline="", encoding="utf-8") as fh:
        writer = csv.writer(fh)
        writer.writerow(HEADERS)
        writer.writerows(BOTS)
    print(f"CSV written  → {output_path}")


def _write_xlsx(output_path: Path) -> None:
    """Write bot data to an interactive Excel workbook."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import (
            Alignment,
            Border,
            Font,
            PatternFill,
            Side,
        )
        from openpyxl.utils import get_column_letter
    except ImportError as exc:
        raise SystemExit(
            "openpyxl is required to generate the Excel file.\n"
            "Install it with:  pip install openpyxl"
        ) from exc

    wb = Workbook()
    ws = wb.active
    ws.title = "DreamCo Bots"

    # ── Header row ──────────────────────────────────────────────────────
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    thin_side = Side(style="thin", color="BFBFBF")
    thin_border = Border(
        left=thin_side, right=thin_side, top=thin_side, bottom=thin_side
    )

    for col_idx, header in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border

    ws.row_dimensions[1].height = 30

    # ── Data rows ────────────────────────────────────────────────────────
    for row_idx, row in enumerate(BOTS, start=2):
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center", wrap_text=(col_idx == 7))

            # Colour-code the ROI Priority column (column 6)
            if col_idx == 6:
                color = ROI_COLORS.get(str(value), "FFFFFF")
                cell.fill = PatternFill(
                    start_color=color, end_color=color, fill_type="solid"
                )
                cell.alignment = Alignment(horizontal="center", vertical="center")

            # Right-align the price column (column 3)
            if col_idx == 3:
                cell.number_format = '"$"#,##0'
                cell.alignment = Alignment(horizontal="right", vertical="center")

        # Alternating row shade for readability
        if row_idx % 2 == 0:
            for col_idx in range(1, len(HEADERS) + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                if col_idx != 6:  # preserve ROI colour
                    cell.fill = PatternFill(
                        start_color="EBF3FB", end_color="EBF3FB", fill_type="solid"
                    )

    # ── Column widths ────────────────────────────────────────────────────
    col_widths = [38, 30, 16, 22, 22, 14, 70]
    for col_idx, width in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # ── AutoFilter ───────────────────────────────────────────────────────
    last_col = get_column_letter(len(HEADERS))
    last_row = len(BOTS) + 1
    ws.auto_filter.ref = f"A1:{last_col}{last_row}"

    # ── Freeze header row ────────────────────────────────────────────────
    ws.freeze_panes = "A2"

    # ── Summary sheet ────────────────────────────────────────────────────
    ws_summary = wb.create_sheet(title="Summary")
    ws_summary["A1"] = "DreamCo Bots — Summary"
    ws_summary["A1"].font = Font(name="Calibri", bold=True, size=14, color="1F4E79")
    ws_summary.column_dimensions["A"].width = 30
    ws_summary.column_dimensions["B"].width = 12

    # Count by industry
    industry_counts: dict[str, int] = {}
    for row in BOTS:
        industry_counts[row[3]] = industry_counts.get(row[3], 0) + 1

    ws_summary["A3"] = "Industry"
    ws_summary["B3"] = "Bot Count"
    ws_summary["A3"].font = Font(bold=True)
    ws_summary["B3"].font = Font(bold=True)
    for sr, (industry, count) in enumerate(sorted(industry_counts.items()), start=4):
        ws_summary.cell(row=sr, column=1, value=industry)
        ws_summary.cell(row=sr, column=2, value=count)

    # Count by ROI priority
    offset = len(industry_counts) + 6
    ws_summary.cell(row=offset, column=1, value="ROI Priority").font = Font(bold=True)
    ws_summary.cell(row=offset, column=2, value="Bot Count").font = Font(bold=True)
    roi_counts: dict[str, int] = {}
    for row in BOTS:
        roi_counts[row[5]] = roi_counts.get(row[5], 0) + 1
    for sr, (priority, count) in enumerate(roi_counts.items(), start=offset + 1):
        cell_p = ws_summary.cell(row=sr, column=1, value=priority)
        cell_c = ws_summary.cell(row=sr, column=2, value=count)
        color = ROI_COLORS.get(priority, "FFFFFF")
        cell_p.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        cell_c.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

    wb.save(output_path)
    print(f"Excel written → {output_path}")


def main() -> None:
    data_dir = Path(__file__).parent
    data_dir.mkdir(parents=True, exist_ok=True)

    csv_path = data_dir / "dreamco_bots_data.csv"
    xlsx_path = data_dir / "dreamco_bots_data.xlsx"

    _write_csv(csv_path)
    _write_xlsx(xlsx_path)

    print(f"\nTotal bots catalogued: {len(BOTS)}")


if __name__ == "__main__":
    main()
